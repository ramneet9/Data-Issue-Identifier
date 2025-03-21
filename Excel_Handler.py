"""Handles Excel file operations including loading, column matching, and coloring."""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from Config import COLORS


def assign_colors(logical, pattern, dtype, colors, priorities):
    """
    Assign colors to cells based on issues and priorities, excluding 'Duplicates'.
    """
    cell_colors = {}
    i = 0
    logical_cols = list(logical.keys())
    while i < len(logical_cols):
        col = logical_cols[i]
        if col != "Duplicates":
            j = 0
            while j < len(logical[col]):
                row = logical[col][j]
                key = (row, col)
                if key not in cell_colors or priorities["logical"] > priorities.get(cell_colors[key][1], 0):
                    cell_colors[key] = (colors["logical"], "logical")
                j += 1
        i += 1
    
    i = 0
    while i < len(pattern):
        col = list(pattern.keys())[i]
        j = 0
        while j < len(pattern[col]):
            row = pattern[col][j]
            key = (row, col)
            if key not in cell_colors or priorities["pattern"] > priorities.get(cell_colors[key][1], 0):
                cell_colors[key] = (colors["pattern"], "pattern")
            j += 1
        i += 1
    
    i = 0
    while i < len(dtype):
        col = list(dtype.keys())[i]
        j = 0
        while j < len(dtype[col]):
            row = dtype[col][j]
            key = (row, col)
            if key not in cell_colors or priorities["dtype"] > priorities.get(cell_colors[key][1], 0):
                cell_colors[key] = (colors["dtype"], "dtype")
            j += 1
        i += 1
    
    return {k: v[0] for k, v in cell_colors.items()}


def apply_colors_to_excel(input_file, cell_colors, column_fill_ratios, output_file, logical_row_issues, pattern_row_issues, dtype_row_issues):
    """
    Applies colors to specific cells, adds Flag and Issues columns, and freezes the first two columns in an Excel file,
    all in memory without creating a temporary file.
    
    Parameters:
        input_file (str): Path to the input Excel file.
        cell_colors (dict): Dictionary where keys are (row, column_name) tuples and values are hex color codes.
        column_fill_ratios (dict): Dictionary where keys are column names and values are fill ratios.
        output_file (str): Path to save the modified Excel file.
        logical_row_issues (dict): Row-wise issues from Logical.py.
        pattern_row_issues (dict): Row-wise issues from pattern.py (per column).
        dtype_row_issues (dict): Row-wise issues from Data_Type.py.
    """
    # Load the original DataFrame
    df = pd.read_excel(input_file)
    
    # Combine all row issues
    all_issues = {}
    for row_idx, issues in logical_row_issues.items():
        all_issues[row_idx] = all_issues.get(row_idx, []) + issues
    for col in pattern_row_issues:
        for row_idx, issues in pattern_row_issues[col].items():
            all_issues[row_idx] = all_issues.get(row_idx, []) + issues
    for row_idx, issues in dtype_row_issues.items():
        all_issues[row_idx] = all_issues.get(row_idx, []) + issues
    
    # Add Flag and Issues columns
    df.insert(0, "Flag", True)  # Default to True (no issues)
    df.insert(1, "Issues", "")
    for row_idx in range(len(df)):
        if row_idx in all_issues:
            df.at[row_idx, "Flag"] = False
            df.at[row_idx, "Issues"] = "; ".join(all_issues[row_idx])
    
    # Create a new workbook in memory
    wb = Workbook()
    ws = wb.active
    
    # Write headers
    all_columns = df.columns.tolist()
    for col_idx, col_name in enumerate(all_columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # Write data
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Adjust cell_colors for column indices (since we added Flag and Issues)
    adjusted_cell_colors = {}
    for (row, col_name), color in cell_colors.items():
        if col_name in all_columns:
            col_idx = all_columns.index(col_name) + 1  # 1-based index
            adjusted_cell_colors[(row, col_idx)] = color
    
    # Apply colors to specific cells
    cell_keys = list(adjusted_cell_colors.keys())
    i = 0
    while i < len(cell_keys):
        (row, col_idx) = cell_keys[i]
        color = adjusted_cell_colors[(row, col_idx)]
        cell = ws.cell(row=row + 2, column=col_idx)  # Adjust for header row
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        i += 1
    
    # Apply color to column headers if fill ratio is below 50%
    column_keys = list(column_fill_ratios.keys())
    j = 0
    while j < len(column_keys):
        col_name = column_keys[j]
        fill_ratio = column_fill_ratios[col_name]
        if fill_ratio < 0.5 and col_name in all_columns:
            col_idx = all_columns.index(col_name) + 1
            header_cell = ws.cell(row=1, column=col_idx)
            header_cell.fill = PatternFill(start_color=COLORS["fill"], end_color=COLORS["fill"], fill_type="solid")
        j += 1
    
    # Freeze the first two columns (Flag and Issues)
    ws.freeze_panes = "C2"  # Freezes columns A and B
    
    # Save the final file
    wb.save(output_file)
    print(f"Saved file with colored cells, headers, and frozen columns: {output_file}")


if __name__ == "__main__":
    input_file = "sample.xlsx"
    output_file = "sample_processed.xlsx"
    logical_row_issues = {0: ["DOB in future"], 1: ["Invalid phone"]}
    pattern_row_issues = {"col1": {0: ["Low coverage pattern"]}, "col2": {2: ["Pattern issue"]}}
    dtype_row_issues = {1: ["Invalid date format"]}
    cell_colors = {(0, "col1"): "ea697e", (1, "col2"): "e1ea69"}
    column_fill_ratios = {"col1": 0.4, "col2": 0.6}
    
    apply_colors_to_excel(input_file, cell_colors, column_fill_ratios, output_file, logical_row_issues, pattern_row_issues, dtype_row_issues)